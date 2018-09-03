from openpyxl import load_workbook
wb = load_workbook("pbb.xlsx")
print(wb.sheetnames)
sheet = wb.get_sheet_by_name("Sheet1")
print(sheet["C"])
print(sheet["B1"].value)
sheet['C3'] = 'Hello world!'
import datetime
sheet['D1']=datetime.datetime.now().strftime("%Y-%m-%d")
wb.save('pbb.xlsx')
